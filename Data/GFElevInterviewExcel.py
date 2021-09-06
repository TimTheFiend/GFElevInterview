
import os
import sys
import sqlite3

import openpyxl

DATABASE_PATH = r"C:\Users\joak\repos\GFElevInterview\bin\Debug\netcoreapp3.1\elevDB.db"
EXCEL_PATH = r"C:\Users\joak\Downloads\Udskrift af data ark. VISI 010621.xlsx"

column_dictionary = {
    "cpr": 1,
    "fornavn": 2,
    "efternavn": 3
}

DB_TABLES = {
    "cpr": "CprNr",
    "fornavn": "Fornavn",
    "efternavn": "Efternavn"
}

def add_to_database(student_list : list):
    os.makedirs("HELLO WORLD HELLO WORLD")
    with sqlite3.connect(DATABASE_PATH) as connection:
        cursor = connection.cursor()
        cursor.executemany("INSERT INTO Elever (CprNr, Fornavn, Efternavn, ElevType) VALUES(?, ?, ?, 0)", student_list)


def main(file_path):
    def get_cell_value(_x, _y):
        return sheet.cell(row=_x + 1, column=_y).value

    student_list = []

    wb = openpyxl.load_workbook(file_path)
    sheet = wb["Udskrift"]

    for x in range(1, sheet.max_row):
        cpr_nummer = get_cell_value(x, column_dictionary['cpr'])

        if cpr_nummer != None:
            student_list.append(
                [
                    cpr_nummer,
                    get_cell_value(x, column_dictionary['fornavn']),
                    get_cell_value(x, column_dictionary['efternavn']),
                ]
            )

    if len(student_list) > 0:
        add_to_database(student_list)


#if len(sys.argv) > 1:
#    main(sys.argv[1])


print("HEHEHEHEHEHHEHEHEH")