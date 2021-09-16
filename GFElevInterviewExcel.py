#!.venv/Scripts/python.exe
"""Forsøger at udprinte indholdet fra et excel-ark hvor filstien bliver givet.

Supplementerende script for `GFElevInterview` projektet, til brug af TEC til grundforløb samtaler.


Joakim Filip Kløigaard Krugstrup.

CHANGELOG:
    10/09-21: Færdigskrevet, tilføjelse (og oversættelse) af kommentarer på dansk.
"""
import sys

## Constants
# Dict: where the key is the error, and the value is the `str` error message that will get sent to C#
ERROR_MESSAGES = {
    'noFilepath': "ERR00: No filepath was given.",
    'emptyStudentList': "ERR01: No students have been found.",
    'columnsMissing': "ERR02: Excel doesn't contain the proper columns.",
    'failedImport': 'ERR03: Failed to import library "openpyxl".',
}
# Dict: Key is a shorthand for the column name, while the value is the entire name.
EXCEL_COLUMN_NAMES = {
    'cpr': 'cprnr',
    'fnavn': 'fornavn',
    'enavn': 'efternavn'
}

COL_POS = {
    'cpr': 0,
    'fnavn': 0,
    'enavn': 0
}

STUDENT_LIST = []
## End Constants.


def print_students() -> None:
    """Prints out the student's info for C# to read."""
    for cpr, fornavn, efternavn in STUDENT_LIST:
        print(f"{cpr};{fornavn};{efternavn};")


def print_error_message(error_msg : str) -> None:
    """Prints out an error message, that's supposed to signal c# to display an error message.

    Args:
        A `str` value, preferably from the `ERROR_MESSAGES` dictionary.
    """
    print(error_msg)


def main(file_path : str) -> None:
    # Check if ´openpyxl´ is available, else exit after displaying an error message.
    try:
        import openpyxl
    except ImportError:
        print_error_message(ERROR_MESSAGES['failedImport'])
        exit()

    def get_cell_value(x : int, y : int) -> str:
        """Returns the value of the cell from the active worksheet.

        Returns:
            The given cell's value as a `str`.
        """
        return SHEET.cell(row=x, column=y).value

    def get_column_positions() -> bool:
        """Reads the first rows' content, and assigns the position to the matching `COLUMN_NAMES` key

        Returns:
            `True` if none of the items in `COLUMN_NAMES` have a value of `0`; else returns `False`
        """
        counter = 1
        while True:
            current_col = get_cell_value(1, counter)

            # Either we run out of columns to look in, or we've set all the values in `COL_POS`
            if current_col == None or 0 not in (COL_POS['cpr'], COL_POS['fnavn'], COL_POS['enavn']):
                break

            if current_col.lower().startswith(EXCEL_COLUMN_NAMES['cpr']):
                COL_POS['cpr'] = counter
            elif current_col.lower().startswith(EXCEL_COLUMN_NAMES['fnavn']):
                COL_POS['fnavn'] = counter
            elif current_col.lower().startswith(EXCEL_COLUMN_NAMES['enavn']):
                COL_POS['enavn'] = counter

            counter += 1

        # Returns true if all columns have a value
        return 0 not in (COL_POS['cpr'], COL_POS['fnavn'], COL_POS['enavn'])


    # Set `WORKBOOK` to the selected file, and set `SHEET` to be the first sheet in the excel file.
    WORKBOOK = openpyxl.load_workbook(file_path)
    SHEET = WORKBOOK[WORKBOOK.sheetnames[0]]

    if get_column_positions() == False:
        print_error_message(ERROR_MESSAGES['columnsMissing'])
        return

    # Read the entire sheet, and add a student to `STUDENT_LIST` if Ax contains a `str`.
    for x in range(2, SHEET.max_row):  # `range(2` since row 1 is the column names
        cpr = get_cell_value(x, COL_POS['cpr'])

        if cpr != None:  # If there's value in the cell, continue.
            STUDENT_LIST.append(
                [
                    cpr,
                    get_cell_value(x, COL_POS['fnavn']),
                    get_cell_value(x, COL_POS['enavn'])
                ]
            )

    # Final check if error
    if len(STUDENT_LIST) > 0:
        print_students()
    else:
        print_error_message(ERROR_MESSAGES['emptyStudentList'])


# Hvis `sys.argv` ikke indeholder 2 elementer, så bliver scriptet ikke kørt.
# sys.argv[0] = dette script, sys.argv[1] = excel filsti
if len(sys.argv) > 1:
    main(sys.argv[1])
else:
    print_error_message(ERROR_MESSAGES['noFilepath'])